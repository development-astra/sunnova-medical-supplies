"""
Sunnova Medical Supplies — Admin dashboard routes (database-backed).

Every screen reads from the database and serializes rows into the dict shapes the
templates already consume. Mutating flows (product CRUD, category create, order
status, inventory adjustments, application review) persist changes, write an
AuditLog row for accountability, and flash feedback. Access is gated by real
login + role-based permissions.
"""

import csv
import io
import secrets
from datetime import datetime, timedelta, timezone
from functools import wraps

from flask import (
    Blueprint, render_template, request, redirect, url_for, abort, flash, current_app, Response,
)
from flask_login import login_user, logout_user, login_required, current_user
from markupsafe import Markup
from sqlalchemy.exc import IntegrityError

from extensions import db
from storage import get_storage, is_allowed, is_image, MAX_BYTES
from emailer import send_email
import admin_data as data
from models import (
    AdminUser, Category, Product, ProductImage, ProductVariant, Customer, Order, OrderItem, Quote, Application,
    InventoryAdjustment, Delivery, Upload, AuditLog, Setting,
    Coupon, Payment, Review, ContentBlock, EmailTemplate, SearchSynonym,
    PromotedProduct, SearchQuery, ProductView, CartAdd, Cart, ContactSubmission,
    ACTION_TONE, ROLE_META, ROLE_PERMISSIONS,
)

admin = Blueprint("admin", __name__, url_prefix="/admin")


# ---------------------------------------------------------------------------
# Helpers: audit logging, RBAC, safe redirects
# ---------------------------------------------------------------------------
def log_action(action, target):
    """Write one audit-trail row for the current admin's action."""
    db.session.add(AuditLog(
        actor=current_user.name if current_user.is_authenticated else "System",
        action=action, target=target,
        ip=(request.headers.get("X-Forwarded-For") or request.remote_addr or "-"),
        tone=ACTION_TONE.get(action, "blue"),
    ))


def role_required(permission):
    """Gate a view behind login + a role permission."""
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("admin.login", next=request.path))
            if not current_user.can(permission):
                abort(403)
            return view(*args, **kwargs)
        return wrapped
    return decorator


def _safe_next(target):
    return target if target and target.startswith("/admin") else url_for("admin.dashboard")


# ---------------------------------------------------------------------------
# Shared list filtering + pagination (used by every admin list screen)
# ---------------------------------------------------------------------------
PER_PAGE = 20


def _arg(name, default=""):
    return (request.args.get(name) or default).strip()


def _like(query, *columns):
    """Apply the ?q= search term across the given columns with ILIKE-style OR."""
    term = _arg("q")
    if not term:
        return query
    from sqlalchemy import or_
    like = f"%{term}%"
    return query.filter(or_(*[c.ilike(like) for c in columns]))


def _date_filtered(query, column):
    """Apply the ?date= range (today/7/30/90 days) to a datetime column."""
    days = {"today": 1, "7": 7, "30": 30, "90": 90}.get(_arg("date"))
    if days:
        cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=days)
        query = query.filter(column >= cutoff)
    return query


def _distinct(column):
    """Sorted distinct non-empty values of a column (for filter dropdowns)."""
    return sorted({v[0] for v in db.session.query(column).distinct().all() if v[0]})


def _paginate(query):
    """Slice a query by ?page= and return (items, meta) for the pagination macro."""
    page = max(_to_int(request.args.get("page"), 1), 1)
    total = query.count()
    pages = max((total + PER_PAGE - 1) // PER_PAGE, 1)
    page = min(page, pages)
    items = query.offset((page - 1) * PER_PAGE).limit(PER_PAGE).all()
    return items, {
        "page": page, "pages": pages, "total": total, "per_page": PER_PAGE,
        "start": ((page - 1) * PER_PAGE + 1) if total else 0,
        "end": min(page * PER_PAGE, total),
        "has_prev": page > 1, "has_next": page < pages,
        "prev": page - 1, "next": page + 1,
    }


# ---------------------------------------------------------------------------
# Template globals (blueprint-scoped): sidebar nav w/ live badges + current admin
# ---------------------------------------------------------------------------
@admin.context_processor
def _inject_admin_globals():
    # Live top-bar notification list (only when signed in — avoids queries on /login).
    notes = _notifications() if current_user.is_authenticated else []
    return {"nav_sections": _build_nav(), "current_admin": _current_admin(),
            "topbar_notifications": notes, "topbar_notification_count": len(notes)}


def _current_admin():
    if current_user.is_authenticated:
        return {"name": current_user.name, "role": current_user.role,
                "email": current_user.email, "initials": current_user.initials}
    return {"name": "Guest", "role": "", "email": "", "initials": "?"}


def _build_nav():
    # Live badge counts (cheap: small tables).
    try:
        low = sum(1 for p in Product.query.all() if p.availability in ("Low Stock", "Out of Stock"))
        badges = {
            "inventory": low,
            "orders": Order.query.filter_by(status="Pending").count(),
            "quotes": Quote.query.filter_by(status="New").count(),
            "delivery": Delivery.query.filter_by(status="Out for Delivery").count(),
            "applications": Application.query.filter(
                Application.status.in_(["New", "Under Review", "More Info Needed"])).count(),
        }
    except Exception:
        badges = {}

    def b(key):
        v = badges.get(key)
        return str(v) if v else None

    return [
        {"heading": None, "items": [
            {"label": "Dashboard", "endpoint": "admin.dashboard", "icon": "grid", "badge": None}]},
        {"heading": "Catalog", "items": [
            {"label": "Products", "endpoint": "admin.products", "icon": "box", "badge": None},
            {"label": "Categories", "endpoint": "admin.categories", "icon": "tag", "badge": None},
            {"label": "Inventory", "endpoint": "admin.inventory", "icon": "layers", "badge": b("inventory")}]},
        {"heading": "Sales", "items": [
            {"label": "Orders", "endpoint": "admin.orders", "icon": "cart", "badge": b("orders")},
            {"label": "Quotes", "endpoint": "admin.quotes", "icon": "file-text", "badge": b("quotes")},
            {"label": "Delivery", "endpoint": "admin.delivery", "icon": "truck", "badge": b("delivery")},
            {"label": "Coupons", "endpoint": "admin.coupons", "icon": "percent", "badge": None},
            {"label": "Payments", "endpoint": "admin.payments", "icon": "credit-card", "badge": None},
            {"label": "Reports", "endpoint": "admin.reports", "icon": "trending-up", "badge": None}]},
        {"heading": "People", "items": [
            {"label": "Customers", "endpoint": "admin.customers", "icon": "users", "badge": None},
            {"label": "Account Applications", "endpoint": "admin.applications", "icon": "user-plus", "badge": b("applications")}]},
        {"heading": "Store", "items": [
            {"label": "Content", "endpoint": "admin.content", "icon": "edit", "badge": None},
            {"label": "Email Templates", "endpoint": "admin.email_templates", "icon": "mail", "badge": None},
            {"label": "Reviews", "endpoint": "admin.reviews", "icon": "star", "badge": None},
            {"label": "Search", "endpoint": "admin.search", "icon": "search", "badge": None}]},
        {"heading": "Operations", "items": [
            {"label": "Uploads", "endpoint": "admin.uploads", "icon": "image", "badge": None},
            {"label": "Notifications", "endpoint": "admin.notifications", "icon": "bell", "badge": None},
            {"label": "Admin Users", "endpoint": "admin.admin_users", "icon": "shield", "badge": None},
            {"label": "Audit Logs", "endpoint": "admin.audit_logs", "icon": "activity", "badge": None},
            {"label": "Settings", "endpoint": "admin.settings", "icon": "settings", "badge": None}]},
    ]


@admin.app_errorhandler(403)
def _forbidden(_e):
    return render_template("admin/403.html"), 403


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
@admin.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("admin.dashboard"))
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = AdminUser.query.filter_by(email=email).first()
        if user and user.status == "Disabled":
            flash("This account has been disabled. Contact a Super Admin.", "error")
        elif user and user.check_password(password):
            login_user(user, remember=bool(request.form.get("remember")))
            user.last_active = "Online now"
            log_action("Admin login", f"{user.role} session")
            db.session.commit()
            return redirect(_safe_next(request.args.get("next")))
        else:
            flash("Incorrect email or password.", "error")
    return render_template("admin/login.html")


@admin.route("/logout")
@login_required
def logout():
    logout_user()
    flash("You have been signed out.", "success")
    return redirect(url_for("admin.login"))


def _now_naive():
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _new_reset_token(user, hours=48):
    """Issue a single-use, time-limited token for invite / password reset links."""
    user.reset_token = secrets.token_urlsafe(24)
    user.reset_expires = _now_naive() + timedelta(hours=hours)
    return user.reset_token


def _password_error(pw, confirm):
    """Enforce the 'require secure password' rule (spec §19)."""
    if pw != confirm:
        return "Passwords do not match."
    if len(pw) < 8:
        return "Use at least 8 characters."
    if not any(c.isdigit() for c in pw) or not any(c.isalpha() for c in pw):
        return "Use both letters and numbers."
    return None


@admin.route("/forgot", methods=["GET", "POST"])
def forgot():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        user = AdminUser.query.filter_by(email=email).first()
        # Always show the same message (don't reveal which emails exist).
        if user and user.status != "Disabled":
            token = _new_reset_token(user)
            send_email(user.email, key="password_reset", context={"name": user.name},
                       cta_text="Reset your password",
                       cta_url=url_for("admin.set_password", token=token, _external=True))
            db.session.commit()
        flash("If that email belongs to an admin, a reset link is on its way.", "success")
        return redirect(url_for("admin.login"))
    return render_template("admin/forgot.html")


@admin.route("/set-password/<token>", methods=["GET", "POST"])
def set_password(token):
    user = AdminUser.query.filter_by(reset_token=token).first()
    if user is None or (user.reset_expires and user.reset_expires < _now_naive()):
        return render_template("admin/set_password.html", invalid=True), 400
    if request.method == "POST":
        pw, confirm = request.form.get("password", ""), request.form.get("confirm", "")
        err = _password_error(pw, confirm)
        if err:
            flash(err, "error")
        else:
            user.set_password(pw)
            user.reset_token, user.reset_expires = None, None
            if user.status == "Invited":
                user.status = "Active"
            db.session.add(AuditLog(actor=user.name, action="Admin login",
                                    target="Password set via secure link", tone="blue"))
            db.session.commit()
            flash("Password set — you can now sign in.", "success")
            return redirect(url_for("admin.login"))
    return render_template("admin/set_password.html", invalid=False, user=user)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
def _as_naive(dt):
    return dt.replace(tzinfo=None) if dt and getattr(dt, "tzinfo", None) else dt


def _sales_series(days=12):
    """Real per-day revenue for the last `days`, as (values, 'Mon DD' labels)."""
    today = datetime.now().date()
    buckets = {today - timedelta(days=days - 1 - i): 0.0 for i in range(days)}
    for o in Order.query.all():
        if o.created_at and o.created_at.date() in buckets:
            buckets[o.created_at.date()] += o.total
    ordered = sorted(buckets.items())
    return [round(v) for _, v in ordered], [d.strftime("%b %d") for d, _ in ordered]


def _best_sellers(limit=5):
    """Top products by revenue, computed from actual order line items."""
    by = {}
    for o in Order.query.all():
        for it in o.items:
            e = by.setdefault(it.sku, {"name": it.name, "sku": it.sku, "units": 0, "rev": 0.0})
            e["units"] += it.qty
            e["rev"] += it.subtotal
    rows = sorted(by.values(), key=lambda e: e["rev"], reverse=True)[:limit]
    out = []
    for r in rows:
        p = Product.query.filter_by(sku=r["sku"]).first()
        out.append({"name": r["name"], "sku": r["sku"], "sold": r["units"],
                    "revenue": f"${r['rev']:,.0f}", "stock": p.stock if p else "—"})
    return out


def _abandoned_cart_count():
    """Active carts with items and no activity in the last hour."""
    cutoff = datetime.now() - timedelta(hours=1)
    return sum(1 for c in Cart.query.filter_by(status="active").all()
               if c.items and c.updated_at and _as_naive(c.updated_at) < cutoff)


@admin.route("/")
@role_required("dashboard")
def dashboard():
    orders = Order.query.all()
    products = Product.query.all()
    today = datetime.now().date()
    total_sales = sum(o.total for o in orders)
    today_sales = sum(o.total for o in orders if o.created_at and o.created_at.date() == today)
    pending = [o for o in orders if o.status == "Pending"]
    low = [p for p in products if p.availability in ("Low Stock", "Out of Stock")]
    out_delivery = Delivery.query.filter_by(status="Out for Delivery").count()
    new_quotes = Quote.query.filter_by(status="New").count()
    open_apps = Application.query.filter(Application.status.in_(["New", "Under Review", "More Info Needed"])).count()
    cutoff30 = datetime.now() - timedelta(days=30)
    new_customers = sum(1 for c in Customer.query.all() if c.created_at and _as_naive(c.created_at) >= cutoff30)
    abandoned = _abandoned_cart_count()

    metrics = [
        {"label": "Total Sales", "value": f"${total_sales:,.0f}", "delta": f"{len(orders)} orders", "trend": "up", "note": "all time", "icon": "dollar", "tone": "brand"},
        {"label": "Today's Sales", "value": f"${today_sales:,.0f}", "delta": "so far today", "trend": "up", "note": today.strftime("%b %d"), "icon": "trending-up", "tone": "green"},
        {"label": "Pending Orders", "value": str(len(pending)), "delta": "needs action", "trend": "flat", "note": "awaiting processing", "icon": "cart", "tone": "amber"},
        {"label": "Pending Quotes", "value": str(new_quotes), "delta": "to review", "trend": "up", "note": "new requests", "icon": "file-text", "tone": "blue"},
        {"label": "Low-Stock Products", "value": str(len(low)), "delta": "reorder soon", "trend": "down", "note": "at/below threshold", "icon": "alert", "tone": "red"},
        {"label": "New Customers", "value": str(new_customers), "delta": "last 30 days", "trend": "up", "note": f"{Customer.query.count()} total", "icon": "users", "tone": "teal"},
        {"label": "Open Applications", "value": str(open_apps), "delta": "to review", "trend": "up", "note": "account requests", "icon": "user-plus", "tone": "blue"},
        {"label": "Out for Delivery", "value": str(out_delivery), "delta": "on route", "trend": "flat", "note": "Miami-Dade", "icon": "truck", "tone": "brand"},
        {"label": "Abandoned Carts", "value": str(abandoned), "delta": "1h+ idle", "trend": "down", "note": "with items", "icon": "cart", "tone": "red"},
    ]

    # Recent activity comes straight from the audit trail now.
    activity = [_activity_from_log(l) for l in AuditLog.query.order_by(AuditLog.created_at.desc()).limit(8).all()]
    sales_trend, sales_labels = _sales_series()

    return render_template(
        "admin/dashboard.html", page_title="Dashboard", metrics=metrics,
        sales_trend=sales_trend, sales_labels=sales_labels,
        best_sellers=_best_sellers(), activity=activity,
    )


_ACTION_ICON = {
    "Admin login": "user", "Product created": "box", "Product updated": "box",
    "Product archived": "archive", "Price changed": "dollar", "Inventory changed": "layers",
    "Order status changed": "cart", "New order placed": "cart", "Quote converted to order": "file-text",
    "Quote status changed": "file-text", "Quote submitted": "file-text",
    "Account approved": "user-plus", "Account rejected": "user-plus", "Account suspended": "user-plus",
    "Password reset": "lock", "Delivery updated": "truck", "Payment retry": "credit-card",
    "Contact form submission": "mail", "Coupon created": "percent",
    "Category created": "tag", "Category updated": "tag", "Category archived": "tag",
    "Settings changed": "settings", "File deleted": "trash",
}


def _activity_from_log(l):
    # Markup("...") % (...) escapes the interpolated values, so a product/customer
    # name containing HTML can't inject script into the dashboard feed (rendered |safe).
    return {
        "icon": _ACTION_ICON.get(l.action, "activity"), "tone": l.tone,
        "text": Markup("%s: <strong>%s</strong>") % (l.action, l.target),
        "time": f"{l.actor} · {l.to_dict()['time']}",
    }


# ---------------------------------------------------------------------------
# Products
# ---------------------------------------------------------------------------
@admin.route("/products")
@role_required("products")
def products():
    q = Product.query
    q = _like(q, Product.name, Product.sku)
    status = _arg("status")
    if status in ("Published", "Draft", "Archived"):
        q = q.filter(Product.status == status)
    if _arg("category"):
        q = q.filter(Product.category == _arg("category"))
    if _arg("brand"):
        q = q.filter(Product.brand == _arg("brand"))
    if _arg("sale_mode"):
        q = q.filter(Product.sale_mode == _arg("sale_mode"))
    stock = _arg("stock")
    if stock == "Out of Stock":
        q = q.filter(Product.stock == 0)
    elif stock == "Low Stock":
        q = q.filter(Product.stock > 0, Product.stock <= Product.threshold)
    elif stock == "In Stock":
        q = q.filter(Product.stock > Product.threshold)
    if _arg("featured") == "1":
        q = q.filter(Product.featured.is_(True))

    items, pg = _paginate(q.order_by(Product.id))
    counts = {
        "all": Product.query.count(),
        "Published": Product.query.filter_by(status="Published").count(),
        "Draft": Product.query.filter_by(status="Draft").count(),
        "Archived": Product.query.filter_by(status="Archived").count(),
    }
    return render_template(
        "admin/products/list.html", page_title="Products",
        products=[p.to_dict() for p in items], pg=pg, counts=counts, filters=request.args,
        categories=[c.to_dict() for c in Category.query.order_by(Category.sort).all()],
        brands=data.BRANDS, sale_modes=data.SALE_MODES,
        stock_options=["In Stock", "Low Stock", "Out of Stock"],
    )


@admin.route("/products/<int:product_id>/duplicate", methods=["POST"])
@role_required("products")
def product_duplicate(product_id):
    p = db.session.get(Product, product_id)
    if p is None:
        abort(404)
    # Generate a unique SKU/slug for the copy.
    base_sku, n = f"{p.sku}-COPY", 1
    new_sku = base_sku
    while Product.query.filter_by(sku=new_sku).first():
        n += 1
        new_sku = f"{base_sku}{n}"
    copy = Product(
        name=f"{p.name} (Copy)", sku=new_sku, category=p.category, brand=p.brand,
        price=p.price, sale_price=p.sale_price, cost=p.cost, stock=0, threshold=p.threshold,
        sale_mode=p.sale_mode, featured=False, status="Draft", short=p.short,
        img_tone=p.img_tone, subcategory=p.subcategory, description=p.description,
        specs=p.specs, seo_title=p.seo_title, seo_description=p.seo_description,
        slug=_unique_product_slug(f"{p.slug or _slugify(p.name)}-copy"))
    db.session.add(copy)
    log_action("Product created", f"{copy.name} ({copy.sku}) — duplicated")
    db.session.commit()
    flash(f"Duplicated “{p.name}” as a draft. Edit and publish when ready.", "success")
    return redirect(url_for("admin.product_edit", product_id=copy.id))


@admin.route("/products/bulk", methods=["POST"])
@role_required("products")
def products_bulk():
    ids = request.form.getlist("ids")
    action = request.form.get("action")
    targets = Product.query.filter(Product.id.in_([_to_int(i, 0) for i in ids])).all()
    if not targets:
        flash("Select at least one product first.", "error")
        return redirect(url_for("admin.products"))
    for p in targets:
        if action == "archive":
            p.status = "Archived"
        elif action == "feature":
            p.featured = True
        elif action == "unfeature":
            p.featured = False
    verb = {"archive": "archived", "feature": "featured", "unfeature": "unfeatured"}.get(action, "updated")
    log_action("Product updated", f"Bulk {verb} {len(targets)} product(s)")
    db.session.commit()
    flash(f"{len(targets)} product(s) {verb}.", "success")
    return redirect(url_for("admin.products"))


def _slugify(text):
    out = "".join(ch.lower() if ch.isalnum() else "-" for ch in (text or ""))
    while "--" in out:
        out = out.replace("--", "-")
    return out.strip("-")


def _unique_product_slug(base, exclude_id=None):
    base = base or "product"
    slug, n = base, 1
    while True:
        existing = Product.query.filter_by(slug=slug).first()
        if existing is None or existing.id == exclude_id:
            return slug
        n += 1
        slug = f"{base}-{n}"


def _apply_product_form(p):
    p.name = request.form.get("name", "").strip()
    p.sku = request.form.get("sku", "").strip()
    p.short = request.form.get("short", "").strip()
    p.category = request.form.get("category") or None
    p.brand = request.form.get("brand") or None
    p.subcategory = request.form.get("subcategory", "").strip()
    p.description = request.form.get("description", "").strip()
    p.specs = request.form.get("specs", "").strip()
    p.seo_title = request.form.get("seo_title", "").strip()
    p.seo_description = request.form.get("seo_description", "").strip()
    p.sale_mode = request.form.get("sale_mode") or "Direct Purchase"
    p.price = _to_float(request.form.get("price"), 0.0)
    p.sale_price = _to_float(request.form.get("sale_price"), None)
    p.cost = _to_float(request.form.get("cost"), None)
    p.stock = _to_int(request.form.get("stock"), 0)
    p.threshold = _to_int(request.form.get("threshold"), 0)
    p.featured = bool(request.form.get("featured"))
    p.status = "Published" if request.form.get("published") else "Draft"
    # Slug: honor a provided value, else derive from the name; keep it unique.
    requested = request.form.get("slug", "").strip() or _slugify(p.name)
    p.slug = _unique_product_slug(_slugify(requested), exclude_id=p.id)


def _save_product_images(p):
    """Save uploaded product images to storage and attach ProductImage rows."""
    files = request.files.getlist("images")
    store = get_storage(current_app)
    saved = 0
    for f in files:
        if not f or not f.filename or not is_image(f.filename):
            continue
        blob = f.read()
        if not blob or len(blob) > MAX_BYTES:
            continue
        key, url = store.save("products", f.filename, blob)
        db.session.add(ProductImage(product_id=p.id, url=url, key=key, sort=len(p.images) + saved))
        if not p.image_url:
            p.image_url = url
        saved += 1
    return saved


@admin.route("/products/new", methods=["GET", "POST"])
@role_required("products")
def product_new():
    if request.method == "POST":
        if not request.form.get("name") or not request.form.get("sku"):
            flash("Product name and SKU are required.", "error")
        else:
            p = Product(img_tone="blue")
            _apply_product_form(p)
            db.session.add(p)
            try:
                db.session.flush()
                _save_product_images(p)
                log_action("Product created", f"{p.name} ({p.sku})")
                db.session.commit()
                flash(f"Product “{p.name}” created.", "success")
                return redirect(url_for("admin.products"))
            except IntegrityError:
                db.session.rollback()
                flash(f"A product with SKU {request.form.get('sku')} already exists.", "error")
    return render_template(
        "admin/products/form.html", page_title="Add Product", product=None,
        categories=[c.to_dict() for c in Category.query.order_by(Category.sort).all()],
        brands=data.BRANDS, sale_modes=data.SALE_MODES, availability=data.AVAILABILITY,
    )


@admin.route("/products/<int:product_id>/edit", methods=["GET", "POST"])
@role_required("products")
def product_edit(product_id):
    p = db.session.get(Product, product_id)
    if p is None:
        abort(404)
    if request.method == "POST":
        old_price = p.price
        _apply_product_form(p)
        try:
            db.session.flush()
            _save_product_images(p)
            log_action("Product updated", f"{p.name} ({p.sku})")
            if old_price != p.price:
                log_action("Price changed", f"{p.name}: ${old_price:.2f} → ${p.price:.2f}")
            db.session.commit()
            flash(f"Product “{p.name}” updated.", "success")
            return redirect(url_for("admin.products"))
        except IntegrityError:
            db.session.rollback()
            flash("Could not save — SKU may already be in use.", "error")
    return render_template(
        "admin/products/form.html", page_title="Edit Product", product=p.to_dict(),
        categories=[c.to_dict() for c in Category.query.order_by(Category.sort).all()],
        brands=data.BRANDS, sale_modes=data.SALE_MODES, availability=data.AVAILABILITY,
    )


@admin.route("/products/<int:product_id>/archive", methods=["POST"])
@role_required("products")
def product_archive(product_id):
    p = db.session.get(Product, product_id)
    if p is None:
        abort(404)
    p.status = "Archived"
    log_action("Product archived", f"{p.name} ({p.sku})")
    db.session.commit()
    flash(f"Product “{p.name}” archived.", "success")
    return redirect(url_for("admin.products"))


@admin.route("/products/<int:product_id>/variant", methods=["POST"])
@role_required("products")
def product_variant_add(product_id):
    p = db.session.get(Product, product_id)
    if p is None:
        abort(404)
    name = request.form.get("variant_name", "").strip()
    if name:
        db.session.add(ProductVariant(
            product_id=p.id, name=name, sku_suffix=request.form.get("variant_sku", "").strip(),
            price_delta=_to_float(request.form.get("variant_price"), 0.0),
            stock=_to_int(request.form.get("variant_stock"), 0)))
        log_action("Product updated", f"{p.name}: variant “{name}” added")
        db.session.commit()
        flash(f"Variant “{name}” added.", "success")
    else:
        flash("Enter a variant name.", "error")
    return redirect(url_for("admin.product_edit", product_id=product_id))


@admin.route("/products/variant/<int:variant_id>/delete", methods=["POST"])
@role_required("products")
def product_variant_delete(variant_id):
    v = db.session.get(ProductVariant, variant_id)
    if v is None:
        abort(404)
    pid = v.product_id
    db.session.delete(v)
    log_action("Product updated", f"Variant removed ({variant_id})")
    db.session.commit()
    flash("Variant removed.", "success")
    return redirect(url_for("admin.product_edit", product_id=pid))


@admin.route("/products/image/<int:image_id>/delete", methods=["POST"])
@role_required("products")
def product_image_delete(image_id):
    im = db.session.get(ProductImage, image_id)
    if im is None:
        abort(404)
    pid = im.product_id
    p = db.session.get(Product, pid)
    get_storage(current_app).delete(im.key)
    was_primary = (p and p.image_url == im.url)
    db.session.delete(im)
    db.session.flush()
    if was_primary:  # promote the next remaining image to primary
        nxt = ProductImage.query.filter_by(product_id=pid).order_by(ProductImage.sort).first()
        p.image_url = nxt.url if nxt else None
    log_action("Product updated", f"Image removed ({image_id})")
    db.session.commit()
    flash("Image removed.", "success")
    return redirect(url_for("admin.product_edit", product_id=pid))


# ---------------------------------------------------------------------------
# Categories
# ---------------------------------------------------------------------------
def _apply_category_form(c):
    c.name = request.form.get("name", "").strip()
    c.slug = request.form.get("slug", "").strip() or _slugify(c.name)
    c.parent = request.form.get("parent") or None
    c.description = request.form.get("description", "")
    c.seo_title = request.form.get("seo_title", "").strip()
    c.seo_description = request.form.get("seo_description", "").strip()
    c.featured = bool(request.form.get("featured"))
    c.sort = _to_int(request.form.get("sort"), 0)
    # Optional category image upload → storage.
    f = request.files.get("image")
    if f and f.filename and is_image(f.filename):
        blob = f.read()
        if blob and len(blob) <= MAX_BYTES:
            key, url = get_storage(current_app).save("categories", f.filename, blob)
            c.image_url = url


@admin.route("/categories", methods=["GET", "POST"])
@role_required("categories")
def categories():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        cat_id = _to_int(request.form.get("category_id"), 0)
        if not name:
            flash("Category name is required.", "error")
        else:
            existing = db.session.get(Category, cat_id) if cat_id else None
            c = existing or Category(status="Published")
            _apply_category_form(c)
            if existing is None:
                db.session.add(c)
            try:
                db.session.flush()
                log_action("Category updated" if existing else "Category created", name)
                db.session.commit()
                flash(f"Category “{name}” {'updated' if existing else 'created'}.", "success")
                return redirect(url_for("admin.categories"))
            except IntegrityError:
                db.session.rollback()
                flash("A category with that slug already exists.", "error")
    q = Category.query
    q = _like(q, Category.name, Category.slug)
    if _arg("status") in ("Published", "Draft", "Archived"):
        q = q.filter(Category.status == _arg("status"))
    if _arg("featured") == "1":
        q = q.filter(Category.featured.is_(True))
    elif _arg("featured") == "0":
        q = q.filter(Category.featured.is_(False))
    return render_template(
        "admin/categories/list.html", page_title="Categories", filters=request.args,
        categories=[c.to_dict() for c in q.order_by(Category.sort).all()],
        all_categories=[c.to_dict() for c in Category.query.order_by(Category.sort).all()],
    )


@admin.route("/categories/<int:category_id>/archive", methods=["POST"])
@role_required("categories")
def category_archive(category_id):
    c = db.session.get(Category, category_id)
    if c is None:
        abort(404)
    # Toggle archive/restore so the row isn't lost.
    c.status = "Published" if c.status == "Archived" else "Archived"
    log_action("Category archived" if c.status == "Archived" else "Category updated",
               f"{c.name} {'archived' if c.status == 'Archived' else 'restored'}")
    db.session.commit()
    flash(f"Category “{c.name}” {'archived' if c.status == 'Archived' else 'restored'}.", "success")
    return redirect(url_for("admin.categories"))


@admin.route("/categories/<int:category_id>/delete", methods=["POST"])
@role_required("categories")
def category_delete(category_id):
    c = db.session.get(Category, category_id)
    if c is None:
        abort(404)
    if c.product_count():
        flash(f"“{c.name}” still has products — reassign or archive it instead.", "error")
        return redirect(url_for("admin.categories"))
    name = c.name
    db.session.delete(c)
    log_action("Category archived", f"{name} deleted")
    db.session.commit()
    flash(f"Category “{name}” deleted.", "success")
    return redirect(url_for("admin.categories"))


# ---------------------------------------------------------------------------
# Orders
# ---------------------------------------------------------------------------
@admin.route("/orders")
@role_required("orders")
def orders():
    q = Order.query
    q = _like(q, Order.id, Order.customer_name, Order.business)
    status = _arg("status")
    if status in data.ORDER_STATUSES:
        q = q.filter(Order.status == status)
    if _arg("payment"):
        q = q.filter(Order.payment == _arg("payment"))
    if _arg("delivery"):
        q = q.filter(Order.delivery == _arg("delivery"))
    q = _date_filtered(q, Order.created_at)
    items, pg = _paginate(q.order_by(Order.created_at.desc()))
    counts = {"all": Order.query.count()}
    for s in data.ORDER_STATUSES:
        counts[s] = Order.query.filter_by(status=s).count()
    return render_template("admin/orders/list.html", page_title="Orders",
                           orders=[o.to_dict() for o in items], pg=pg, counts=counts, filters=request.args,
                           order_statuses=data.ORDER_STATUSES, payment_statuses=data.PAYMENT_STATUSES,
                           delivery_methods=["Local Delivery", "Pickup"])


@admin.route("/orders/<order_id>")
@role_required("orders")
def order_detail(order_id):
    o = db.session.get(Order, order_id)
    if o is None:
        abort(404)
    return render_template("admin/orders/detail.html", page_title="Order " + order_id,
                           order=o.to_detail_dict(), order_statuses=data.ORDER_STATUSES)


@admin.route("/orders/<order_id>/status", methods=["POST"])
@role_required("orders")
def order_status(order_id):
    o = db.session.get(Order, order_id)
    if o is None:
        abort(404)
    new_status = request.form.get("status", o.status)
    if new_status != o.status and new_status in data.ORDER_STATUSES:
        old = o.status
        o.status = new_status
        log_action("Order status changed", f"{o.id}: {old} → {new_status}")
        # Keep the customer informed of the change.
        send_email(o.email, key="order_status",
                   context={"name": o.customer_name, "order_id": o.id, "status": new_status})
        db.session.commit()
        flash(f"Order {o.id} marked {new_status}. Customer emailed.", "success")
    return redirect(url_for("admin.order_detail", order_id=order_id))


@admin.route("/orders/<order_id>/email", methods=["POST"])
@role_required("orders")
def order_email(order_id):
    o = db.session.get(Order, order_id)
    if o is None:
        abort(404)
    sent = send_email(o.email, key="order_status",
                      context={"name": o.customer_name, "order_id": o.id, "status": o.status})
    flash(f"Order update email sent to {o.email}." if sent else "No email on file for this order.",
          "success" if sent else "error")
    return redirect(url_for("admin.order_detail", order_id=order_id))


@admin.route("/orders/<order_id>/note", methods=["POST"])
@role_required("orders")
def order_note(order_id):
    o = db.session.get(Order, order_id)
    if o is None:
        abort(404)
    o.internal_notes = request.form.get("internal_notes", o.internal_notes)
    log_action("Order status changed", f"{o.id}: internal note updated")
    db.session.commit()
    flash("Internal note saved.", "success")
    return redirect(url_for("admin.order_detail", order_id=order_id))


@admin.route("/orders/<order_id>/pay", methods=["POST"])
@role_required("orders")
def order_pay(order_id):
    o = db.session.get(Order, order_id)
    if o is None:
        abort(404)
    o.payment = "Paid"
    pay = Payment.query.filter_by(order_id=o.id).first()
    if pay:
        pay.status, pay.refund_status = "Paid", "—"
    log_action("Order status changed", f"{o.id} payment marked Paid")
    db.session.commit()
    flash(f"Order {o.id} marked paid.", "success")
    return redirect(url_for("admin.order_detail", order_id=order_id))


@admin.route("/orders/<order_id>/print/<doc>")
@role_required("orders")
def order_print(order_id, doc):
    o = db.session.get(Order, order_id)
    if o is None or doc not in ("invoice", "packing-slip"):
        abort(404)
    return render_template("admin/orders/print.html", order=o.to_detail_dict(),
                           doc=doc, store=_settings_dict().get("store", {}))


# ---------------------------------------------------------------------------
# Quotes
# ---------------------------------------------------------------------------
@admin.route("/quotes")
@role_required("quotes")
def quotes():
    query = Quote.query
    query = _like(query, Quote.id, Quote.customer_name, Quote.business)
    status = _arg("status")
    if status in data.QUOTE_STATUSES:
        query = query.filter(Quote.status == status)
    if _arg("assigned"):
        query = query.filter(Quote.assigned == _arg("assigned"))
    items, pg = _paginate(query.order_by(Quote.created_at.desc()))
    counts = {"all": Quote.query.count()}
    for s in data.QUOTE_STATUSES:
        counts[s] = Quote.query.filter_by(status=s).count()
    return render_template("admin/quotes/list.html", page_title="Quotes",
                           quotes=[q.to_dict() for q in items], pg=pg, counts=counts, filters=request.args,
                           quote_statuses=data.QUOTE_STATUSES, assignees=_distinct(Quote.assigned))


@admin.route("/quotes/<quote_id>")
@role_required("quotes")
def quote_detail(quote_id):
    q = db.session.get(Quote, quote_id)
    if q is None:
        abort(404)
    return render_template("admin/quotes/detail.html", page_title="Quote " + quote_id,
                           quote=q.to_detail_dict(), quote_statuses=data.QUOTE_STATUSES)


@admin.route("/quotes/<quote_id>/status", methods=["POST"])
@role_required("quotes")
def quote_status(quote_id):
    q = db.session.get(Quote, quote_id)
    if q is None:
        abort(404)
    new = request.form.get("status", q.status)
    if new != q.status and new in data.QUOTE_STATUSES:
        old, q.status = q.status, new
        log_action("Quote status changed", f"{q.id}: {old} → {new}")
        db.session.commit()
        flash(f"Quote {q.id} marked {new}.", "success")
    return redirect(url_for("admin.quote_detail", quote_id=quote_id))


@admin.route("/quotes/<quote_id>/note", methods=["POST"])
@role_required("quotes")
def quote_note(quote_id):
    q = db.session.get(Quote, quote_id)
    if q is None:
        abort(404)
    q.internal_notes = request.form.get("internal_notes", q.internal_notes)
    log_action("Quote status changed", f"{q.id}: internal note updated")
    db.session.commit()
    flash("Internal note saved.", "success")
    return redirect(url_for("admin.quote_detail", quote_id=quote_id))


@admin.route("/quotes/<quote_id>/assign", methods=["POST"])
@role_required("quotes")
def quote_assign(quote_id):
    q = db.session.get(Quote, quote_id)
    if q is None:
        abort(404)
    q.assigned = request.form.get("assigned", q.assigned)
    log_action("Quote status changed", f"{q.id} assigned to {q.assigned}")
    db.session.commit()
    flash(f"Quote {q.id} assigned to {q.assigned}.", "success")
    return redirect(url_for("admin.quote_detail", quote_id=quote_id))


@admin.route("/quotes/<quote_id>/quote", methods=["POST"])
@role_required("quotes")
def quote_price(quote_id):
    q = db.session.get(Quote, quote_id)
    if q is None:
        abort(404)
    total = 0.0
    for item in q.items:
        val = _to_float(request.form.get("quoted_" + item.sku), None)
        item.quoted = val
        if val:
            total += val * item.qty
    q.value = f"${total:,.2f}" if total else "-"
    q.status = "Responded"
    if request.form.get("internal_notes") is not None:
        q.internal_notes = request.form.get("internal_notes")
    log_action("Quote status changed", f"{q.id}: prices quoted, marked Responded")
    send_email(q.email, key="quote_response", context={"name": q.customer_name, "quote_id": q.id},
               cta_text="View your quote", cta_url=url_for("index", _external=True))
    db.session.commit()
    flash(f"Quoted prices saved for {q.id}. Quote marked Responded and customer emailed.", "success")
    return redirect(url_for("admin.quote_detail", quote_id=quote_id))


@admin.route("/quotes/<quote_id>/convert", methods=["POST"])
@role_required("quotes")
def quote_convert(quote_id):
    q = db.session.get(Quote, quote_id)
    if q is None:
        abort(404)
    if q.status == "Converted to Order":
        flash(f"Quote {q.id} was already converted.", "warning")
        return redirect(url_for("admin.quote_detail", quote_id=quote_id))
    new_id = _next_order_id()
    order = Order(
        id=new_id, customer_id=q.customer_id, customer_name=q.customer_name,
        business=q.business, email=q.email, phone=q.phone, customer_type=q.customer_type,
        delivery_address=f"{q.business}, Miami-Dade County, FL",
        billing_address=f"{q.business}, Miami-Dade County, FL",
        payment="Pending", status="Pending", delivery="Local Delivery",
        delivery_status="Awaiting Scheduling")
    for item in q.items:
        price = item.quoted
        if price is None:
            prod = Product.query.filter_by(sku=item.sku).first()
            price = (prod.sale_price or prod.price) if prod else 0.0
        order.items.append(OrderItem(name=item.name, sku=item.sku, qty=item.qty, price=price))
    db.session.add(order)
    q.status = "Converted to Order"
    log_action("Quote converted to order", f"{q.id} → {new_id}")
    db.session.commit()
    flash(f"Quote {q.id} converted to order {new_id}.", "success")
    return redirect(url_for("admin.order_detail", order_id=new_id))


def _next_order_id():
    nums = []
    for o in Order.query.all():
        if o.id.startswith("SN-") and o.id.split("-")[1].isdigit():
            nums.append(int(o.id.split("-")[1]))
    return f"SN-{(max(nums) + 1) if nums else 10001}"


# ---------------------------------------------------------------------------
# Customers
# ---------------------------------------------------------------------------
@admin.route("/customers")
@role_required("customers")
def customers():
    q = Customer.query
    q = _like(q, Customer.name, Customer.email, Customer.business, Customer.phone)
    ctype = _arg("type")
    if ctype in data.CUSTOMER_TYPES:
        q = q.filter(Customer.type == ctype)
    if _arg("account"):
        q = q.filter(Customer.account == _arg("account"))
    items, pg = _paginate(q.order_by(Customer.name))
    counts = {"all": Customer.query.count()}
    for t in data.CUSTOMER_TYPES:
        counts[t] = Customer.query.filter_by(type=t).count()
    return render_template("admin/customers/list.html", page_title="Customers",
                           customers=[c.to_dict() for c in items], pg=pg, counts=counts, filters=request.args,
                           customer_types=data.CUSTOMER_TYPES,
                           account_statuses=["Verified", "Pending", "Suspended"])


@admin.route("/customers/<int:customer_id>")
@role_required("customers")
def customer_detail(customer_id):
    c = db.session.get(Customer, customer_id)
    if c is None:
        abort(404)
    return render_template("admin/customers/detail.html", page_title=c.name,
                           customer=_customer_detail_dict(c))


def _customer_detail_dict(c):
    orders = Order.query.filter_by(customer_id=c.id).order_by(Order.created_at.desc()).all()
    quotes = Quote.query.filter_by(customer_id=c.id).order_by(Quote.created_at.desc()).all()
    spent = sum(o.total for o in orders)
    addr = data.CUSTOMER_DETAIL["addresses"] if c.id == data.CUSTOMER_DETAIL["id"] else \
        [{"label": "Delivery", "value": f"{c.business or c.name}, Miami-Dade County, FL"}]
    docs = data.CUSTOMER_DETAIL["documents"] if c.id == data.CUSTOMER_DETAIL["id"] else []
    return {
        "id": c.id, "name": c.name, "email": c.email, "phone": c.phone,
        "business": c.business or "-", "type": c.type, "account": c.account,
        "member_since": c.member_since or "-", "orders": len(orders), "spent": spent,
        "avg_order": (spent / len(orders)) if orders else 0.0,
        "last_order": c.last_order_date(), "addresses": addr,
        "recent_orders": [{"id": o.id, "date": o.to_dict()["date"], "total": o.total, "status": o.status} for o in orders[:5]],
        "quotes": [{"id": q.id, "date": q.to_dict()["date"], "status": q.status, "value": q.value} for q in quotes],
        "documents": docs, "notes": c.notes or "No internal notes yet.",
    }


@admin.route("/customers/<int:customer_id>/action", methods=["POST"])
@role_required("customers")
def customer_action(customer_id):
    c = db.session.get(Customer, customer_id)
    if c is None:
        abort(404)
    action = request.form.get("action")
    if action == "approve":
        c.account = "Verified"
        log_action("Account approved", f"{c.name} ({c.business or 'individual'})")
        flash(f"{c.name}'s account approved.", "success")
    elif action == "suspend":
        c.account = "Suspended"
        log_action("Account suspended", f"{c.name} ({c.business or 'individual'})")
        flash(f"{c.name}'s account suspended.", "success")
    elif action == "reset":
        sent = send_email(c.email, key="password_reset", context={"name": c.name},
                          cta_text="Reset your password", cta_url=url_for("index", _external=True))
        log_action("Password reset", f"Reset email {'sent to' if sent else 'requested for'} {c.email}")
        flash(f"Password reset email sent to {c.email}." if sent
              else f"No email on file for {c.name}.", "success" if sent else "error")
    elif action == "note":
        c.notes = request.form.get("note", c.notes)
        flash("Internal note saved.", "success")
    db.session.commit()
    return redirect(url_for("admin.customer_detail", customer_id=customer_id))


# ---------------------------------------------------------------------------
# Account Applications
# ---------------------------------------------------------------------------
@admin.route("/applications")
@role_required("applications")
def applications():
    q = Application.query
    q = _like(q, Application.business, Application.contact, Application.email)
    status = _arg("status")
    if status in data.APPLICATION_STATUSES:
        q = q.filter(Application.status == status)
    if _arg("type"):
        q = q.filter(Application.type == _arg("type"))
    items, pg = _paginate(q.order_by(Application.created_at.desc()))
    counts = {"all": Application.query.count()}
    for s in data.APPLICATION_STATUSES:
        counts[s] = Application.query.filter_by(status=s).count()
    return render_template("admin/applications/list.html", page_title="Account Applications",
                           applications=[a.to_dict() for a in items], pg=pg, counts=counts, filters=request.args,
                           application_statuses=data.APPLICATION_STATUSES, customer_types=data.CUSTOMER_TYPES)


@admin.route("/applications/<app_id>/action", methods=["POST"])
@role_required("applications")
def application_action(app_id):
    a = db.session.get(Application, app_id)
    if a is None:
        abort(404)
    action = request.form.get("action")
    mapping = {"approve": ("Approved", "Account approved"), "reject": ("Rejected", "Account rejected"),
               "info": ("More Info Needed", "Account application updated")}
    if action in mapping:
        a.status, audit = mapping[action]
        if request.form.get("note"):
            a.notes = request.form.get("note")
        # Approving a business account promotes it into the customer database.
        if action == "approve" and not Customer.query.filter_by(email=a.email).first():
            db.session.add(Customer(name=a.contact, email=a.email, phone=a.phone,
                                    business=a.business, type=a.type, account="Verified",
                                    member_since="2026"))
        # Notify the applicant of the decision.
        if action == "approve":
            send_email(a.email, key="account_approved", context={"name": a.contact},
                       cta_text="Start ordering", cta_url=url_for("index", _external=True))
        elif action == "reject":
            send_email(a.email, subject="Update on your Sunnova account application",
                       context={"name": a.contact},
                       body=(f"Hi {a.contact},\n\nThank you for your interest in a Sunnova business "
                             f"account. After review we're unable to approve the application at this "
                             f"time. Please contact us if you'd like to discuss."))
        elif action == "info":
            send_email(a.email, subject="More information needed for your Sunnova application",
                       context={"name": a.contact},
                       body=(f"Hi {a.contact},\n\nWe need a little more information to finish reviewing "
                             f"your business account application. Please reply with your practice details "
                             f"and any licensing documents."))
        log_action(audit, f"{a.business} ({a.id})")
        db.session.commit()
        flash(f"Application {a.id} — {a.status.lower()}. Applicant notified by email.", "success")
    return redirect(url_for("admin.applications"))


# ---------------------------------------------------------------------------
# Inventory
# ---------------------------------------------------------------------------
@admin.route("/inventory")
@role_required("inventory")
def inventory():
    q = Product.query
    q = _like(q, Product.name, Product.sku)
    stock = _arg("stock")
    if stock == "Out of Stock":
        q = q.filter(Product.stock == 0)
    elif stock == "Low Stock":
        q = q.filter(Product.stock > 0, Product.stock <= Product.threshold)
    elif stock == "In Stock":
        q = q.filter(Product.stock > Product.threshold)
    items, pg = _paginate(q.order_by(Product.id))
    # Stock-health tiles must reflect the WHOLE catalog, not just the current
    # page — compute them from every product using the same availability logic.
    stock_counts = {"Low Stock": 0, "Out of Stock": 0, "In Stock": 0}
    for p in Product.query.all():
        av = p.availability
        if av in stock_counts:
            stock_counts[av] += 1
    hist = [h.to_dict() for h in InventoryAdjustment.query.order_by(InventoryAdjustment.created_at.desc()).limit(12).all()]
    return render_template("admin/inventory/list.html", page_title="Inventory",
                           inventory=[p.to_inventory_dict() for p in items], pg=pg, filters=request.args,
                           history=hist, reasons=data.ADJUSTMENT_REASONS, stock_counts=stock_counts,
                           stock_options=["In Stock", "Low Stock", "Out of Stock"])


@admin.route("/inventory/adjust", methods=["POST"])
@role_required("inventory")
def inventory_adjust():
    sku = request.form.get("sku")
    reason = request.form.get("reason", "Correct stock count")
    qty = _to_int(request.form.get("quantity"), 0)
    p = Product.query.filter_by(sku=sku).first()
    if p is None or qty == 0:
        flash("Select a product and a non-zero quantity.", "error")
        return redirect(url_for("admin.inventory"))

    # "Reduce"/"damaged"/"returned" remove stock; "Add"/"restock" add; "Correct" sets absolute.
    if reason in ("Reduce stock", "Mark as damaged"):
        change = -abs(qty)
    elif reason == "Correct stock count":
        change = qty - p.stock
    else:  # Add stock, Mark as returned
        change = abs(qty)

    p.stock = max(p.stock + change, 0)
    db.session.add(InventoryAdjustment(sku=p.sku, name=p.name, change=change, reason=reason,
                                       by=current_user.name))
    log_action("Inventory changed", f"{p.name} {'+' if change >= 0 else ''}{change} ({reason})")
    db.session.commit()
    flash(f"Stock adjusted for {p.name} ({'+' if change >= 0 else ''}{change}).", "success")
    return redirect(url_for("admin.inventory"))


# ---------------------------------------------------------------------------
# Delivery
# ---------------------------------------------------------------------------
@admin.route("/delivery")
@role_required("delivery")
def delivery():
    q = Delivery.query
    q = _like(q, Delivery.order_id, Delivery.customer, Delivery.zip, Delivery.address)
    status = _arg("status")
    if status in data.DELIVERY_STATUSES:
        q = q.filter(Delivery.status == status)
    items, pg = _paginate(q.order_by(Delivery.id))
    counts = {"all": Delivery.query.count()}
    for s in data.DELIVERY_STATUSES:
        counts[s] = Delivery.query.filter_by(status=s).count()
    return render_template("admin/delivery/list.html", page_title="Delivery",
                           deliveries=[d.to_dict() for d in items], pg=pg, counts=counts, filters=request.args,
                           delivery_statuses=data.DELIVERY_STATUSES)


@admin.route("/delivery/<int:delivery_id>/action", methods=["POST"])
@role_required("delivery")
def delivery_action(delivery_id):
    d = db.session.get(Delivery, delivery_id)
    if d is None:
        abort(404)
    action = request.form.get("action")
    mapping = {"out": "Out for Delivery", "delivered": "Delivered",
               "reschedule": "Rescheduled", "preparing": "Preparing"}
    if action in mapping:
        d.status = mapping[action]
        if request.form.get("note"):
            d.notes = request.form.get("note")
        log_action("Delivery updated", f"{d.order_id} → {d.status}")
        db.session.commit()
        flash(f"Delivery {d.order_id} marked {d.status}.", "success")
    elif action == "notify":
        order = db.session.get(Order, d.order_id)
        to = order.email if order else None
        sent = send_email(to, key="delivery_update",
                          context={"name": d.customer, "order_id": d.order_id})
        flash(f"Delivery update email sent for {d.order_id}." if sent
              else f"No email on file for {d.order_id}.", "success" if sent else "error")
    return redirect(url_for("admin.delivery"))


@admin.route("/delivery/schedule", methods=["POST"])
@role_required("delivery")
def delivery_schedule():
    d = db.session.get(Delivery, _to_int(request.form.get("delivery_id"), 0))
    if d is None:
        flash("Select a delivery to schedule.", "error")
        return redirect(url_for("admin.delivery"))
    d.date = request.form.get("date") or d.date
    d.driver = request.form.get("driver") or d.driver
    if request.form.get("note"):
        d.notes = request.form.get("note")
    d.status = "Scheduled"
    # Mirror the scheduled date onto the linked order so it shows in the orders list.
    order = db.session.get(Order, d.order_id)
    if order:
        order.delivery_date = d.date
        order.delivery_status = "Scheduled"
    log_action("Delivery updated", f"{d.order_id} scheduled for {d.date}")
    db.session.commit()
    flash(f"Delivery {d.order_id} scheduled for {d.date}.", "success")
    return redirect(url_for("admin.delivery"))


# ---------------------------------------------------------------------------
# Uploads
# ---------------------------------------------------------------------------
@admin.route("/uploads")
@role_required("uploads")
def uploads():
    ups = Upload.query.order_by(Upload.created_at.desc()).all()
    store = get_storage(current_app)
    stats = {
        "total_files": Upload.query.count(), "total_size": data.UPLOAD_STATS["total_size"],
        "images": sum(1 for u in ups if u.is_image), "documents": sum(1 for u in ups if not u.is_image),
        "cdn": (store.cdn.replace("https://", "") if store.kind == "spaces" else "local · static/uploads"),
    }
    return render_template("admin/uploads/index.html", page_title="Uploads",
                           uploads=[u.to_dict() for u in ups], stats=stats,
                           folders=["products", "categories", "quotes", "documents"])


_UPLOAD_TYPE = {
    "products": "Product Image", "categories": "Category Image",
    "quotes": "Quote Attachment", "documents": "Account Document",
}
_TONES = ["blue", "teal", "amber", "green", "purple", "brand"]


@admin.route("/uploads", methods=["POST"])
@role_required("uploads")
def upload_files():
    folder = request.form.get("folder", "products")
    if folder not in _UPLOAD_TYPE:
        folder = "products"
    files = request.files.getlist("files")
    store = get_storage(current_app)
    saved, skipped = 0, 0
    for f in files:
        if not f or not f.filename:
            continue
        if not is_allowed(f.filename):
            skipped += 1
            continue
        blob = f.read()
        if len(blob) > MAX_BYTES:
            skipped += 1
            continue
        key, url = store.save(folder, f.filename, blob)
        img = is_image(f.filename)
        db.session.add(Upload(
            name=f.filename, type=_UPLOAD_TYPE[folder], size=_human_size(len(blob)),
            dims="—", folder=folder, by=current_user.name, tone=_TONES[saved % len(_TONES)],
            url=url, key=key, is_image=img))
        saved += 1
    if saved:
        log_action("Settings changed", f"Uploaded {saved} file(s) to /{folder}")
        db.session.commit()
        flash(f"Uploaded {saved} file(s)." + (f" {skipped} skipped (type/size)." if skipped else ""), "success")
    else:
        flash("No files uploaded. Allowed: PNG, JPG, WEBP, GIF, PDF up to 5MB.", "error")
    return redirect(url_for("admin.uploads"))


@admin.route("/uploads/<int:upload_id>/delete", methods=["POST"])
@role_required("uploads")
def upload_delete(upload_id):
    u = db.session.get(Upload, upload_id)
    if u is None:
        abort(404)
    get_storage(current_app).delete(u.key)
    name = u.name
    db.session.delete(u)
    log_action("File deleted", name)
    db.session.commit()
    flash(f"Deleted {name}.", "success")
    return redirect(url_for("admin.uploads"))


def _human_size(n):
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.0f} KB"
    return f"{n / 1024 / 1024:.1f} MB"


# ---------------------------------------------------------------------------
# Admin Users & Roles
# ---------------------------------------------------------------------------
@admin.route("/admin-users")
@role_required("admin_users")
def admin_users():
    users = [u.to_dict() for u in AdminUser.query.order_by(AdminUser.id).all()]
    roles = [{"name": name, "access": meta["access"], "tone": meta["tone"],
              "users": AdminUser.query.filter_by(role=name).count()}
             for name, meta in ROLE_META.items()]
    return render_template("admin/admin_users/list.html", page_title="Admin Users",
                           users=users, roles=roles)


@admin.route("/admin-users/invite", methods=["POST"])
@role_required("admin_users")
def admin_user_invite():
    email = request.form.get("email", "").strip().lower()
    role = request.form.get("role", "Support Staff")
    name = request.form.get("name", "").strip() or email.split("@")[0].replace(".", " ").title()
    if not email:
        flash("An email address is required to invite an admin.", "error")
    elif AdminUser.query.filter_by(email=email).first():
        flash("An admin with that email already exists.", "error")
    elif role not in ROLE_PERMISSIONS:
        flash("Choose a valid role.", "error")
    else:
        u = AdminUser(name=name, email=email, role=role, status="Invited", last_active="Never")
        u.set_password(secrets.token_urlsafe(16))  # random; the invitee sets their own via the link
        db.session.add(u)
        db.session.flush()
        token = _new_reset_token(u)
        send_email(email, subject=f"You're invited to the Sunnova admin ({role})",
                   body=(f"Hi {name},\n\nYou've been invited to the Sunnova Medical Supplies "
                         f"admin dashboard as {role}. Use the button below to set your password "
                         f"and sign in. This link expires in 48 hours."),
                   cta_text="Set your password",
                   cta_url=url_for("admin.set_password", token=token, _external=True))
        log_action("Settings changed", f"Invited admin {email} as {role}")
        db.session.commit()
        flash(f"Invitation email sent to {email} ({role}).", "success")
    return redirect(url_for("admin.admin_users"))


@admin.route("/admin-users/<int:user_id>/role", methods=["POST"])
@role_required("admin_users")
def admin_user_role(user_id):
    u = db.session.get(AdminUser, user_id)
    if u is None:
        abort(404)
    role = request.form.get("role")
    if role in ROLE_PERMISSIONS and role != u.role:
        old, u.role = u.role, role
        log_action("Settings changed", f"{u.name} role: {old} → {role}")
        db.session.commit()
        flash(f"{u.name}'s role changed to {role}.", "success")
    return redirect(url_for("admin.admin_users"))


@admin.route("/admin-users/<int:user_id>/toggle", methods=["POST"])
@role_required("admin_users")
def admin_user_toggle(user_id):
    u = db.session.get(AdminUser, user_id)
    if u is None:
        abort(404)
    if u.id == current_user.id:
        flash("You can’t disable your own account.", "error")
        return redirect(url_for("admin.admin_users"))
    u.status = "Active" if u.status == "Disabled" else "Disabled"
    log_action("Settings changed", f"{u.name} {'enabled' if u.status == 'Active' else 'disabled'}")
    db.session.commit()
    flash(f"{u.name} {'enabled' if u.status == 'Active' else 'disabled'}.", "success")
    return redirect(url_for("admin.admin_users"))


# ---------------------------------------------------------------------------
# Audit Logs
# ---------------------------------------------------------------------------
@admin.route("/audit-logs")
@role_required("audit_logs")
def audit_logs():
    q = AuditLog.query
    q = _like(q, AuditLog.actor, AuditLog.action, AuditLog.target)
    if _arg("action"):
        q = q.filter(AuditLog.action == _arg("action"))
    q = _date_filtered(q, AuditLog.created_at)
    items, pg = _paginate(q.order_by(AuditLog.created_at.desc()))
    return render_template("admin/audit_logs/list.html", page_title="Audit Logs",
                           logs=[l.to_dict() for l in items], pg=pg, filters=request.args,
                           actions=_distinct(AuditLog.action))


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------
@admin.route("/settings", methods=["GET", "POST"])
@role_required("settings")
def settings():
    if request.method == "POST":
        # form field name -> setting key
        text_fields = {
            "store_name": "store.name", "store_phone": "store.phone",
            "store_support_email": "store.support_email", "store_address": "store.address",
            "store_hours": "store.hours", "store_min_order": "store.min_order",
            "store_free_delivery_threshold": "store.free_delivery_threshold",
            "delivery_fee": "delivery.fee", "delivery_zip_codes": "delivery.zip_codes",
            "delivery_cutoff": "delivery.cutoff", "delivery_days": "delivery.days",
            "social_facebook": "social.facebook", "social_instagram": "social.instagram",
            "social_linkedin": "social.linkedin",
        }
        for field, key in text_fields.items():
            _set_setting(key, request.form.get(field, ""))
        _set_setting("delivery.enabled", "true" if request.form.get("delivery_enabled") else "false")
        _set_setting("delivery.same_week", "true" if request.form.get("delivery_same_week") else "false")
        log_action("Settings changed", "Store settings updated")
        db.session.commit()
        flash("Settings saved.", "success")
        return redirect(url_for("admin.settings"))
    return render_template("admin/settings/index.html", page_title="Settings", settings=_settings_dict())


def _settings_dict():
    out = {"store": {}, "delivery": {}, "social": {}}
    for r in Setting.query.all():
        if "." not in r.key:
            continue
        sec, k = r.key.split(".", 1)
        val = r.value
        if val in ("true", "false"):
            val = (val == "true")
        out.setdefault(sec, {})[k] = val
    return out


def _set_setting(key, value):
    row = db.session.get(Setting, key)
    if row is None:
        row = Setting(key=key)
        db.session.add(row)
    row.value = value


# ---------------------------------------------------------------------------
# Coupons & Discounts
# ---------------------------------------------------------------------------
@admin.route("/coupons")
@role_required("coupons")
def coupons():
    q = Coupon.query
    q = _like(q, Coupon.code, Coupon.scope)
    if _arg("type"):
        q = q.filter(Coupon.type == _arg("type"))
    active = _arg("active")
    if active == "1":
        q = q.filter(Coupon.active.is_(True))
    elif active == "0":
        q = q.filter(Coupon.active.is_(False))
    items, pg = _paginate(q.order_by(Coupon.created_at.desc()))
    return render_template("admin/coupons/list.html", page_title="Coupons",
                           coupons=[c.to_dict() for c in items], pg=pg, filters=request.args,
                           coupon_types=["Percentage", "Fixed", "Free Delivery"])


@admin.route("/coupons/new", methods=["POST"])
@role_required("coupons")
def coupon_new():
    code = request.form.get("code", "").strip().upper()
    if not code:
        flash("A coupon code is required.", "error")
    elif Coupon.query.filter_by(code=code).first():
        flash(f"Coupon {code} already exists.", "error")
    else:
        db.session.add(Coupon(
            code=code, type=request.form.get("type", "Percentage"),
            amount=_to_float(request.form.get("amount"), 0.0),
            min_order=_to_float(request.form.get("min_order"), 0.0),
            usage_limit=_to_int(request.form.get("usage_limit"), 0),
            scope=request.form.get("scope", "All products"),
            expires=request.form.get("expires", ""), active=True))
        log_action("Coupon created", code)
        db.session.commit()
        flash(f"Coupon {code} created.", "success")
    return redirect(url_for("admin.coupons"))


@admin.route("/coupons/<int:coupon_id>/toggle", methods=["POST"])
@role_required("coupons")
def coupon_toggle(coupon_id):
    c = db.session.get(Coupon, coupon_id)
    if c is None:
        abort(404)
    c.active = not c.active
    log_action("Settings changed", f"Coupon {c.code} {'enabled' if c.active else 'disabled'}")
    db.session.commit()
    flash(f"Coupon {c.code} {'enabled' if c.active else 'disabled'}.", "success")
    return redirect(url_for("admin.coupons"))


# ---------------------------------------------------------------------------
# Payments & Transactions
# ---------------------------------------------------------------------------
@admin.route("/payments")
@role_required("payments")
def payments():
    q = Payment.query
    q = _like(q, Payment.order_id, Payment.customer, Payment.txn_id)
    if _arg("status"):
        q = q.filter(Payment.status == _arg("status"))
    if _arg("provider"):
        q = q.filter(Payment.provider == _arg("provider"))
    items, pg = _paginate(q.order_by(Payment.created_at.desc()))
    return render_template("admin/payments/list.html", page_title="Payments",
                           payments=[p.to_dict() for p in items], pg=pg, filters=request.args,
                           payment_statuses=data.PAYMENT_STATUSES, providers=_distinct(Payment.provider))


@admin.route("/payments/<int:payment_id>/action", methods=["POST"])
@role_required("payments")
def payment_action(payment_id):
    p = db.session.get(Payment, payment_id)
    if p is None:
        abort(404)
    action = request.form.get("action")
    if action == "mark_paid":
        p.status = "Paid"
        order = db.session.get(Order, p.order_id)
        if order:
            order.payment = "Paid"
        log_action("Order status changed", f"Payment {p.txn_id} marked Paid")
        flash(f"Payment for {p.order_id} marked paid.", "success")
    elif action == "refund":
        p.status = "Refunded"
        p.refund_status = "Refunded"
        order = db.session.get(Order, p.order_id)
        if order:
            order.payment = "Refunded"
        log_action("Order status changed", f"Payment {p.txn_id} refunded")
        flash(f"Refund issued for {p.order_id}.", "success")
    elif action == "retry":
        order = db.session.get(Order, p.order_id)
        to = order.email if order else None
        sent = send_email(to, subject="Payment needed for your Sunnova order",
                          context={"name": p.customer, "order_id": p.order_id},
                          body=(f"Hi {p.customer},\n\nWe weren't able to process payment for order "
                                f"{p.order_id}. Please retry your payment so we can schedule delivery."),
                          cta_text="Retry payment", cta_url=url_for("index", _external=True))
        log_action("Payment retry", f"{p.order_id} ({p.txn_id})")
        flash(f"Payment retry notification sent for {p.order_id}." if sent
              else f"No email on file for {p.order_id}.", "success" if sent else "error")
    db.session.commit()
    return redirect(url_for("admin.payments"))


# ---------------------------------------------------------------------------
# Reports & Analytics
# ---------------------------------------------------------------------------
def _report_data():
    orders = Order.query.all()
    products = Product.query.all()
    revenue = sum(o.total for o in orders)
    order_count = len(orders)
    aov = revenue / order_count if order_count else 0.0

    # Sales by category (join order items -> product category)
    cat_of = {p.sku: p.category for p in products}
    by_category, by_product = {}, {}
    for o in orders:
        for it in o.items:
            line = it.subtotal
            by_category[cat_of.get(it.sku, "Other")] = by_category.get(cat_of.get(it.sku, "Other"), 0) + line
            by_product[it.name] = by_product.get(it.name, 0) + line

    best = sorted(by_product.items(), key=lambda kv: kv[1], reverse=True)[:5]
    low = [p for p in products if p.availability in ("Low Stock", "Out of Stock")]

    # Revenue by customer type
    by_type = {}
    for o in orders:
        by_type[o.customer_type or "Business"] = by_type.get(o.customer_type or "Business", 0) + o.total

    # Repeat customers
    counts = {}
    for o in orders:
        if o.customer_id:
            counts[o.customer_id] = counts.get(o.customer_id, 0) + 1
    repeat = sum(1 for v in counts.values() if v > 1)

    quotes = Quote.query.all()
    converted = sum(1 for q in quotes if q.status == "Converted to Order")
    conv_rate = (converted / len(quotes) * 100) if quotes else 0

    peak_cat = max((v for _, v in by_category.items()), default=1) or 1

    # Customer lifetime value + top customers by spend
    spends = sorted(((c, c.total_spent()) for c in Customer.query.all()), key=lambda kv: kv[1], reverse=True)
    clv = (sum(s for _, s in spends) / len(spends)) if spends else 0.0
    top_customers = [{"name": c.name, "business": c.business or "-", "orders": c.order_count(), "spent": s}
                     for c, s in spends[:6] if s > 0]

    # Abandoned carts
    cutoff = datetime.now() - timedelta(hours=1)
    abandoned = [c for c in Cart.query.filter_by(status="active").all()
                 if c.items and c.updated_at and _as_naive(c.updated_at) < cutoff]

    # Frequently viewed / added-to-cart (storefront analytics)
    def _tally(rows):
        agg = {}
        for r in rows:
            e = agg.setdefault(r.sku, {"name": r.name, "sku": r.sku, "count": 0})
            e["count"] += 1
        return sorted(agg.values(), key=lambda e: e["count"], reverse=True)[:6]

    return {
        "revenue": revenue, "order_count": order_count, "aov": aov,
        "by_category": [{"name": k, "revenue": v, "pct": round(v / peak_cat * 100)} for k, v in sorted(by_category.items(), key=lambda kv: kv[1], reverse=True)],
        "best": [{"name": k, "revenue": v} for k, v in best],
        "low": [{"name": p.name, "sku": p.sku, "stock": p.stock, "threshold": p.threshold, "availability": p.availability} for p in low],
        "by_type": [{"type": k, "revenue": v} for k, v in sorted(by_type.items(), key=lambda kv: kv[1], reverse=True)],
        "repeat": repeat, "customers": Customer.query.count(),
        "conv_rate": conv_rate, "quote_count": len(quotes),
        "delivery_count": Delivery.query.count(),
        "clv": clv, "top_customers": top_customers,
        "abandoned_count": len(abandoned), "abandoned_value": sum(c.subtotal for c in abandoned),
        "freq_viewed": _tally(ProductView.query.all()),
        "freq_added": _tally(CartAdd.query.all()),
    }


@admin.route("/reports")
@role_required("reports")
def reports():
    return render_template("admin/reports/index.html", page_title="Reports", r=_report_data(),
                           sales_trend=data.SALES_TREND, sales_labels=data.SALES_TREND_LABELS)


@admin.route("/reports/export.csv")
@role_required("reports")
def reports_export():
    r = _report_data()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Sunnova Medical Supplies — Sales Report"])
    w.writerow([])
    w.writerow(["Metric", "Value"])
    w.writerow(["Total revenue", f"{r['revenue']:.2f}"])
    w.writerow(["Orders", r["order_count"]])
    w.writerow(["Average order value", f"{r['aov']:.2f}"])
    w.writerow(["Customers", r["customers"]])
    w.writerow(["Repeat customers", r["repeat"]])
    w.writerow(["Quote conversion rate", f"{r['conv_rate']:.1f}%"])
    w.writerow([])
    w.writerow(["Sales by category", "Revenue"])
    for row in r["by_category"]:
        w.writerow([row["name"], f"{row['revenue']:.2f}"])
    w.writerow([])
    w.writerow(["Revenue by customer type", "Revenue"])
    for row in r["by_type"]:
        w.writerow([row["type"], f"{row['revenue']:.2f}"])
    log_action("Settings changed", "Exported sales report (CSV)")
    db.session.commit()
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=sunnova-sales-report.csv"})


@admin.route("/reports/export.xlsx")
@role_required("reports")
def reports_export_xlsx():
    from openpyxl import Workbook
    r = _report_data()
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Sunnova Medical Supplies — Sales Report"])
    ws.append([])
    ws.append(["Metric", "Value"])
    for label, val in [
        ("Total revenue", round(r["revenue"], 2)), ("Orders", r["order_count"]),
        ("Average order value", round(r["aov"], 2)), ("Customer lifetime value", round(r["clv"], 2)),
        ("Customers", r["customers"]), ("Repeat customers", r["repeat"]),
        ("Quote conversion rate %", round(r["conv_rate"], 1)),
        ("Deliveries", r["delivery_count"]), ("Abandoned carts", r["abandoned_count"]),
    ]:
        ws.append([label, val])
    cat = wb.create_sheet("By category")
    cat.append(["Category", "Revenue"])
    for row in r["by_category"]:
        cat.append([row["name"], round(row["revenue"], 2)])
    typ = wb.create_sheet("By customer type")
    typ.append(["Type", "Revenue"])
    for row in r["by_type"]:
        typ.append([row["type"], round(row["revenue"], 2)])
    best = wb.create_sheet("Best sellers")
    best.append(["Product", "Revenue"])
    for row in r["best"]:
        best.append([row["name"], round(row["revenue"], 2)])
    buf = io.BytesIO()
    wb.save(buf)
    log_action("Settings changed", "Exported sales report (Excel)")
    db.session.commit()
    return Response(buf.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=sunnova-sales-report.xlsx"})


@admin.route("/reports/print")
@role_required("reports")
def reports_print():
    return render_template("admin/reports/print.html", r=_report_data(),
                           store=_settings_dict().get("store", {}))


# ---------------------------------------------------------------------------
# Search Management
# ---------------------------------------------------------------------------
def _popular_searches():
    """Aggregate the raw SearchQuery log into popular-search rows (term, count,
    latest result-count, conversion %). Cross-DB safe (aggregates in Python)."""
    agg = {}
    for q in SearchQuery.query.order_by(SearchQuery.created_at).all():
        d = agg.setdefault(q.term, {"term": q.term, "count": 0, "converted": 0, "results": q.results})
        d["count"] += 1
        d["results"] = q.results  # most recent result count for the term
        if q.converted:
            d["converted"] += 1
    rows = []
    for d in agg.values():
        conv = (d["converted"] / d["count"] * 100) if d["count"] else 0
        rows.append({"term": d["term"], "count": d["count"], "results": d["results"],
                     "conversion": f"{conv:.1f}%", "no_results": d["results"] == 0})
    return sorted(rows, key=lambda r: r["count"], reverse=True)


@admin.route("/search")
@role_required("search")
def search():
    return render_template(
        "admin/search/index.html", page_title="Search",
        synonyms=[s.to_dict() for s in SearchSynonym.query.all()],
        promoted=[p.to_dict() for p in PromotedProduct.query.all()],
        popular=_popular_searches(),
        products=[p.to_dict() for p in Product.query.order_by(Product.name).all()])


@admin.route("/search/synonym", methods=["POST"])
@role_required("search")
def search_synonym():
    term = request.form.get("term", "").strip()
    syns = request.form.get("synonyms", "").strip()
    if term and syns:
        db.session.add(SearchSynonym(term=term, synonyms=syns))
        log_action("Settings changed", f"Search synonym added: {term}")
        db.session.commit()
        flash(f"Synonyms added for “{term}”.", "success")
    else:
        flash("Enter a term and its synonyms.", "error")
    return redirect(url_for("admin.search"))


@admin.route("/search/promote", methods=["POST"])
@role_required("search")
def search_promote():
    keyword = request.form.get("keyword", "").strip()
    sku = request.form.get("sku", "").strip()
    prod = Product.query.filter_by(sku=sku).first()
    if keyword and prod:
        db.session.add(PromotedProduct(keyword=keyword, sku=sku, product_name=prod.name))
        log_action("Settings changed", f"Promoted {sku} for “{keyword}”")
        db.session.commit()
        flash(f"Promoted {prod.name} for “{keyword}”.", "success")
    else:
        flash("Enter a keyword and pick a product.", "error")
    return redirect(url_for("admin.search"))


# ---------------------------------------------------------------------------
# Reviews / Testimonials
# ---------------------------------------------------------------------------
@admin.route("/reviews")
@role_required("reviews")
def reviews():
    items = [r.to_dict() for r in Review.query.order_by(Review.created_at.desc()).all()]
    return render_template("admin/reviews/list.html", page_title="Reviews", reviews=items)


@admin.route("/reviews/<int:review_id>/action", methods=["POST"])
@role_required("reviews")
def review_action(review_id):
    rv = db.session.get(Review, review_id)
    if rv is None:
        abort(404)
    action = request.form.get("action")
    if action == "approve":
        rv.status = "Approved"
        flash("Review approved.", "success")
    elif action == "reject":
        rv.status = "Rejected"
        flash("Review rejected.", "success")
    elif action == "feature":
        rv.featured = not rv.featured
        flash("Review " + ("featured on homepage." if rv.featured else "unfeatured."), "success")
    elif action == "delete":
        db.session.delete(rv)
        flash("Review deleted.", "success")
    log_action("Settings changed", f"Review {action} — {rv.author if action != 'delete' else 'removed'}")
    db.session.commit()
    return redirect(url_for("admin.reviews"))


# ---------------------------------------------------------------------------
# Content management (CMS)
# ---------------------------------------------------------------------------
@admin.route("/content", methods=["GET", "POST"])
@role_required("content")
def content():
    blocks = ContentBlock.query.all()
    if request.method == "POST":
        for b in blocks:
            field = "block_" + b.key.replace(".", "__")
            if field in request.form:
                b.value = request.form.get(field, "")
        log_action("Settings changed", "Website content updated")
        db.session.commit()
        flash("Content saved.", "success")
        return redirect(url_for("admin.content"))
    # group blocks for the template
    groups = {}
    for b in blocks:
        groups.setdefault(b.group, []).append(b.to_dict() | {"field": "block_" + b.key.replace(".", "__")})
    return render_template("admin/content/index.html", page_title="Content", groups=groups)


# ---------------------------------------------------------------------------
# Email templates
# ---------------------------------------------------------------------------
@admin.route("/email-templates")
@role_required("email_templates")
def email_templates():
    items = [t.to_dict() for t in EmailTemplate.query.all()]
    return render_template("admin/email_templates/list.html", page_title="Email Templates", templates=items)


@admin.route("/email-templates/<key>", methods=["POST"])
@role_required("email_templates")
def email_template_save(key):
    t = db.session.get(EmailTemplate, key)
    if t is None:
        abort(404)
    if "toggle" in request.form:
        t.enabled = not t.enabled
        flash(f"{t.name} {'enabled' if t.enabled else 'disabled'}.", "success")
    else:
        t.subject = request.form.get("subject", t.subject)
        t.body = request.form.get("body", t.body)
        flash(f"{t.name} saved.", "success")
    log_action("Settings changed", f"Email template updated: {t.name}")
    db.session.commit()
    return redirect(url_for("admin.email_templates"))


# ---------------------------------------------------------------------------
# Notifications (derived from live state)
# ---------------------------------------------------------------------------
def _notifications():
    notes = []
    pending = Order.query.filter_by(status="Pending").count()
    if pending:
        notes.append({"icon": "cart", "tone": "amber", "title": f"{pending} pending order(s)",
                      "text": "Orders awaiting processing.", "url": url_for("admin.orders")})
    low = [p for p in Product.query.all() if p.availability in ("Low Stock", "Out of Stock")]
    if low:
        notes.append({"icon": "alert", "tone": "red", "title": f"{len(low)} product(s) low/out of stock",
                      "text": "Reorder soon to avoid stockouts.", "url": url_for("admin.inventory")})
    new_q = Quote.query.filter_by(status="New").count()
    if new_q:
        notes.append({"icon": "file-text", "tone": "blue", "title": f"{new_q} new quote request(s)",
                      "text": "Quotes waiting for a response.", "url": url_for("admin.quotes")})
    apps = Application.query.filter(Application.status.in_(["New", "Under Review"])).count()
    if apps:
        notes.append({"icon": "user-plus", "tone": "teal", "title": f"{apps} account application(s)",
                      "text": "New business accounts to review.", "url": url_for("admin.applications")})
    failed = Payment.query.filter_by(status="Failed").count()
    if failed:
        notes.append({"icon": "credit-card", "tone": "red", "title": f"{failed} failed payment(s)",
                      "text": "Payments that need attention.", "url": url_for("admin.payments")})
    out = Delivery.query.filter_by(status="Out for Delivery").count()
    if out:
        notes.append({"icon": "truck", "tone": "brand", "title": f"{out} order(s) out for delivery",
                      "text": "Deliveries currently on route.", "url": url_for("admin.delivery")})
    return notes


@admin.route("/notifications")
@role_required("notifications")
def notifications():
    return render_template("admin/notifications/index.html", page_title="Notifications",
                           notifications=_notifications())


# ---------------------------------------------------------------------------
# Small parse helpers
# ---------------------------------------------------------------------------
def _to_float(value, default):
    try:
        return float(value) if value not in (None, "") else default
    except (ValueError, TypeError):
        return default


def _to_int(value, default):
    try:
        return int(float(value)) if value not in (None, "") else default
    except (ValueError, TypeError):
        return default
